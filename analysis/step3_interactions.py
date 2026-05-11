"""
Step 3 — 2-way interaction tables.
Output: output/interaction_tables.xlsx (one sheet per pair, with
        conditional color formatting).

Pairs (per spec):
  - short_delta × spread_type
  - short_delta × VIX_regime
  - spread_type × VIX_regime
  - p_quartile × win_rate (calibration)
  - G_quartile × avg_dollar_pnl (signal ranking)
  - IV_regime × VIX_regime
"""
import math
import os
import sys
import warnings

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from _helpers import (
    load_trades, load_vix_daily, spy_monday_gap_pct, per_ticker_iv_pctile,
    OUTPUT_DIR,
)
from step2_buckets import (
    assign_short_delta, assign_quartile, assign_vix_regime,
    assign_iv_pctile, assign_gap,
)

OUT_XLSX = os.path.join(OUTPUT_DIR, "interaction_tables.xlsx")
N_FLOOR  = 15

HEADER_FILL = PatternFill(start_color="0F0F0F", end_color="0F0F0F", fill_type="solid")
HEADER_FONT = Font(bold=True, color="E8E8E8")
INSUFFICIENT_FILL = PatternFill(start_color="2A2520", end_color="2A2520", fill_type="solid")


def prep_trades():
    t = load_trades()
    vix = load_vix_daily()[["Date", "Close"]].rename(
        columns={"Date": "entry_date", "Close": "vix_close"}
    )
    t = pd.merge_asof(
        t.sort_values("entry_date"), vix.sort_values("entry_date"),
        on="entry_date", direction="backward",
        tolerance=pd.Timedelta(days=4),
    )
    gap = spy_monday_gap_pct(t["entry_date"])
    t["gap_pct"] = t["entry_date"].map(gap)
    iv = per_ticker_iv_pctile(window_weeks=52, winsor_q=0.99)
    t = t.merge(
        iv.rename(columns={"Symbol": "ticker", "DataDate": "entry_date"}),
        on=["ticker", "entry_date"], how="left"
    )
    t["bk_short_delta"]  = assign_short_delta(t)
    t["bk_p_q"], _       = assign_quartile(t["p"],  "p ")
    t["bk_G_q"], _       = assign_quartile(t["G"],  "G ")
    t["bk_vix"]          = assign_vix_regime(t["vix_close"])
    t["bk_iv_pctile"]    = assign_iv_pctile(t["iv_pctile"])
    return t


def crosstab_metric(df, row_dim, col_dim, agg, metric_name):
    """
    Build a (row_dim × col_dim) pivot table where each cell is `agg(sub_df)`.
    Cells with n < N_FLOOR become NaN. Returns (pivot, n_pivot).
    """
    # Drop rows missing either dim
    sub = df.dropna(subset=[row_dim, col_dim])
    n_pivot = sub.groupby([row_dim, col_dim], observed=True).size().unstack(fill_value=0)
    if metric_name == "n":
        return n_pivot, n_pivot

    metric_pivot = pd.DataFrame(index=n_pivot.index, columns=n_pivot.columns, dtype=float)
    for r in n_pivot.index:
        for c in n_pivot.columns:
            cell = sub[(sub[row_dim] == r) & (sub[col_dim] == c)]
            if len(cell) < N_FLOOR:
                metric_pivot.loc[r, c] = np.nan
            else:
                metric_pivot.loc[r, c] = agg(cell)
    return metric_pivot, n_pivot


def write_section(ws, start_row, title, df_pivot, n_pivot, fmt="0.000",
                  color_rule=None):
    """Write one labeled cross-tab block."""
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    r0 = start_row + 1

    # Column headers
    for j, col in enumerate(df_pivot.columns):
        c = ws.cell(row=r0, column=j + 2, value=str(col))
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    # Row headers + values
    for i, idx in enumerate(df_pivot.index):
        rr = r0 + i + 1
        rc = ws.cell(row=rr, column=1, value=str(idx))
        rc.font = HEADER_FONT
        rc.fill = HEADER_FILL
        for j, col in enumerate(df_pivot.columns):
            val   = df_pivot.loc[idx, col]
            n_val = int(n_pivot.loc[idx, col])
            cell  = ws.cell(row=rr, column=j + 2)
            if pd.isna(val):
                cell.value = f"ins (n={n_val})" if n_val > 0 else ""
                cell.fill  = INSUFFICIENT_FILL
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value = float(val)
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")

    end_row = r0 + len(df_pivot.index)
    end_col = len(df_pivot.columns) + 1
    if color_rule is not None and len(df_pivot.index) > 0:
        rng = f"{get_column_letter(2)}{r0+1}:" \
              f"{get_column_letter(end_col)}{end_row}"
        ws.conditional_formatting.add(rng, color_rule)
    return end_row + 2


def add_interaction_sheet(wb, name, df, row_dim, col_dim,
                          row_label, col_label, n_caption=""):
    ws = wb.create_sheet(title=name[:31])  # sheet names ≤ 31 chars
    ws.column_dimensions["A"].width = 22
    for j in range(2, 30):
        ws.column_dimensions[get_column_letter(j)].width = 14

    title = f"{row_label} × {col_label}"
    if n_caption:
        title += f"   [{n_caption}]"
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=14)

    win_rule = ColorScaleRule(
        start_type="num", start_value=0.50, start_color="E24B4A",
        mid_type="num",   mid_value=0.60,   mid_color="FFFFFF",
        end_type="num",   end_value=0.70,   end_color="1D9E75",
    )
    sharpe_rule = ColorScaleRule(
        start_type="num", start_value=0.0, start_color="E24B4A",
        mid_type="num",   mid_value=0.5,  mid_color="FFFFFF",
        end_type="num",   end_value=1.5,  end_color="1D9E75",
    )
    pnl_rule = ColorScaleRule(
        start_type="num", start_value=-20, start_color="E24B4A",
        mid_type="num",   mid_value=0,    mid_color="FFFFFF",
        end_type="num",   end_value=30,   end_color="1D9E75",
    )

    row = 3
    n_piv, _ = crosstab_metric(df, row_dim, col_dim,
                               agg=lambda s: len(s), metric_name="n")
    row = write_section(ws, row, "n (sample size per cell)",
                        n_piv, n_piv, fmt="0", color_rule=None)

    win_piv, n_piv = crosstab_metric(df, row_dim, col_dim,
                                     agg=lambda s: float(s["is_win"].mean()),
                                     metric_name="win_rate")
    row = write_section(ws, row, "win_rate", win_piv, n_piv,
                        fmt="0.0%", color_rule=win_rule)

    pnl_piv, _ = crosstab_metric(df, row_dim, col_dim,
                                 agg=lambda s: float(s["dollar_pnl"].mean()),
                                 metric_name="avg_dollar_pnl")
    row = write_section(ws, row, "avg_dollar_pnl", pnl_piv, n_piv,
                        fmt='"$"#,##0.00', color_rule=pnl_rule)

    def _sharpe(s):
        std = s["dollar_pnl"].std(ddof=1)
        if std <= 0 or len(s) < 2:
            return float("nan")
        return float(s["dollar_pnl"].mean() / std * math.sqrt(52))

    sh_piv, _ = crosstab_metric(df, row_dim, col_dim, agg=_sharpe,
                                metric_name="sharpe")
    row = write_section(ws, row, "sharpe (annualized × √52, approximate)",
                        sh_piv, n_piv, fmt="0.00", color_rule=sharpe_rule)


def main():
    df = prep_trades()
    print(f"Loaded {len(df):,} trades")

    wb = Workbook()
    # Drop default sheet
    default = wb.active
    wb.remove(default)

    pairs = [
        ("delta×type",       "bk_short_delta", "spread_type",
         "short_delta",      "spread_type"),
        ("delta×VIX",        "bk_short_delta", "bk_vix",
         "short_delta",      "VIX_regime"),
        ("type×VIX",         "spread_type",    "bk_vix",
         "spread_type",      "VIX_regime"),
        ("p_q×type",         "bk_p_q",         "spread_type",
         "p_quartile",       "spread_type"),
        ("G_q×type",         "bk_G_q",         "spread_type",
         "G_quartile",       "spread_type"),
        ("IV_pct×VIX",       "bk_iv_pctile",   "bk_vix",
         "IV_pctile_q",      "VIX_regime"),
    ]

    for name, rd, cd, rl, cl in pairs:
        n_caption = f"floor n≥{N_FLOOR} marked 'ins'"
        add_interaction_sheet(wb, name, df, rd, cd, rl, cl, n_caption)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"\nWrote {OUT_XLSX}")
    print(f"  {len(pairs)} sheets, conditional formatting applied")


if __name__ == "__main__":
    main()
